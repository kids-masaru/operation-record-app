import openpyxl
from copy import copy
from datetime import datetime

def copy_row_style_and_formulas(ws, source_row_idx, target_row_idx):
    """
    Copy values (if formula), styles, and number formats from source row to target row.
    """
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row_idx, column=col)
        target_cell = ws.cell(row=target_row_idx, column=col)
        
        # Copy Style
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            
        # Copy Formula logic: Simple relative adjustment is hard. 
        # For now, just copy the value or formula string exact.
        # Ideally we'd translate the formula reference, but that's complex.
        # If the template implies dragging down, it usually works if we write usage logic.
        target_cell.value = source_cell.value


# Column Mapping
# Based on inspection of sample.xlsx and App 218 Field Codes
# Sheet: 保育園情報
NURSERY_SHEET_NAME = "保育園情報"
NURSERY_MAP = {
    "name": "D",           # 施設名 (Code: name)
    "client_name": "C",    # クライアント名 (Code: client_name)
    "capacity": "E",       # 定員 (Code: capacity)
    "open_date": "F",      # 開園日 (Code: open_date)
    "addr_area": "T",      # 都道府県 (Code: addr_area)
    "addr_city": "U",      # 住所/市区郡 (Code: addr_city)
    # W, X are protected/formulas
}
NURSERY_KEY_COL = 4 # Column D

# Sheet: 病床数
BED_SHEET_NAME = "病床数"
BED_MAP = {
    "保育園": "D",          # 施設名 (Code: 保育園)
    "病床数合計_0": "F",    # 病床数 (Code: 病床数合計_0)
}
BED_KEY_COL = 4 # Column D

def get_column_index(col_letter):
    """Convert 'A'->1, 'B'->2"""
    return openpyxl.utils.column_index_from_string(col_letter)

from copy import copy

def copy_cell_style(source_cell, target_cell):
    """Copy font, border, fill, number_format, protection, alignment."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def find_header_with_text(ws, text_part):
    """Find cell containing text in first 50 columns, 5 rows."""
    for r in range(1, 6):
        for c in range(1, 55):
            val = ws.cell(row=r, column=c).value
            if val and text_part in str(val):
                return ws.cell(row=r, column=c)
    return None

def update_sheet(ws, records, mapping, key_field_kintone="name", key_col_idx=4, start_row=5):
    """
    Generic update function with Style Copying and Insert Row logic.
    """
    # 1. Identify Existing Rows and Last Data Row
    excel_rows = {}
    last_data_row = start_row - 1
    
    # Scan from start_row down to find table end
    MAX_SCAN = 2000
    for row in range(start_row, MAX_SCAN):
        cell_val = ws.cell(row=row, column=key_col_idx).value
        
        if cell_val:
            excel_rows[str(cell_val).strip()] = row
            last_data_row = row
        else:
            # Check if next few rows are empty to confirm end
            # (Checking 3 rows should be enough to avoid false positives in sparse tables)
            if all(not ws.cell(row=r, column=key_col_idx).value for r in range(row+1, row+4)):
                break
            
    # 2. Process Records
    for rec in records:
        key_val = rec.get(key_field_kintone, {}).get("value", "")
        if not key_val: continue
        
        target_row = excel_rows.get(str(key_val).strip())
        
        # [NEW ROW LOGIC]
        if not target_row:
            # Insert after the last known data row
            # Ensure we are at least at start_row
            insert_at = max(last_data_row + 1, start_row)
            ws.insert_rows(insert_at)
            target_row = insert_at
            
            # Copy Styles/Formulas from the row above (if it's a valid data row)
            source_row = insert_at - 1
            # Only copy if source_row is within our data table range (>= start_row)
            # OR if it is the template row (start_row) and we are inserting at start_row + 1?
            # Actually, simply copying from the row above is usually safe in a table.
            # But we must avoid copying the HEADER row (start_row - 1).
            # If insert_at == start_row (empty table?), copying start_row-1 might be header.
            # Safety: If source_row < start_row, maybe don't copy, or warn? 
            # Assuming template has at least 1 row of data or valid formulas at start_row.
            
            if source_row >= start_row or (source_row == start_row - 1 and start_row > 1): 
                # Allow copying from header row IF user wants style, but formulas might be wrong.
                # BETTER: Only copy if source_row >= start_row (meaning we copy an existing sibling data row).
                # If table is empty (last_data_row < start_row), we might have an issue.
                # Assume template has data.
                if source_row >= start_row:
                    for col in range(1, ws.max_column + 1):
                        source_cell = ws.cell(row=source_row, column=col)
                        target_cell = ws.cell(row=target_row, column=col)
                        
                        copy_cell_style(source_cell, target_cell)
                        
                        if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
                            try:
                                translated_formula = Translator(source_cell.value, source_cell.coordinate).translate_formula(target_cell.coordinate)
                                target_cell.value = translated_formula
                            except:
                                target_cell.value = source_cell.value
            
            # Update our tracker
            last_data_row = target_row
            excel_rows[str(key_val).strip()] = target_row

        # [WRITE VALUES]
        for k_field, col_letter in mapping.items():
            if col_letter in ["W", "X"]: continue # PROTECTED
            
            val = rec.get(k_field, {}).get("value", "")
            col_idx = get_column_index(col_letter)
            
            # Type Conversion
            if "日" in k_field and val:
                try:
                    dt = datetime.strptime(val, "%Y-%m-%d")
                    ws.cell(row=target_row, column=col_idx).value = dt
                    ws.cell(row=target_row, column=col_idx).number_format = "yyyy年m月d日"
                    continue
                except: pass
            elif ("数" in k_field or "定員" in k_field) and val:
                try:
                    val = int(val)
                except: pass
            
            ws.cell(row=target_row, column=col_idx).value = val

    # 3. Update Date Header (Once per sheet if relevant)
    # Search for "現在" in top rows
    date_cell = find_header_with_text(ws, "現在")
    if date_cell:
         # Replace date part logic? Or just overwrite "20xx年x月 現在"
         # Assuming user wants "YYYY年M月xD日 現在" or similar
         # config_date is a datetime object
         new_str = config_date.strftime("%Y年%m月 現在") # Guessing format from manual
         date_cell.value = new_str


            

def update_excel(template_file, merged_data, config_date):
    """
    Main function to update the Excel workbook.
    """
    wb = openpyxl.load_workbook(template_file) # data_only=False to keep formulas
    
    # Extract lists
    nursery_list = [m['master'] for m in merged_data]
    # bed_list/client_list comes from the 2nd dataset
    client_list = [m['bed'] for m in merged_data if m['bed']] # Naming 'bed' is from prev code, reused as 'client_data'
    
    sheet_names = wb.sheetnames
    
from openpyxl.formula.translate import Translator

# Sheet 1: クライアント名あり（縦）
# This is the presentation sheet.
CLIENT_SHEET_NAME = "クライアント名あり（縦）"
# Only write the Key (AB). Visible columns (A-Z) are likely formulas referencing AB.
CLIENT_VIEW_MAP = {
    "name": "AB",          # 施設名 (Unique Key)
    # "client_name": "M",  # If M is not a formula, we could write it. But safe to assume lookup.
    # "open_date": "AU",   # If AU is data?
}
CLIENT_VIEW_KEY_COL = 28 # Column AB

def update_excel(template_file, merged_data, config_date):
    """
    Update the first sheet of the workbook with a clean summary list.
    """
    wb = openpyxl.load_workbook(template_file, keep_vba=True)
    
    # Target: First Sheet
    ws = wb.worksheets[0]
    ws.title = "Kintoneデータ抽出"
    
    # Clear existing data (keep row 1 if valuable? No, user wants specific headers)
    # Let's overwrite from A1
    
    # 1. Define Headers
    headers = [
        "住所", "ステータス", "施設名", "クライアント名", "開園日", "基本開園日", "定員", 
        "病児保育", "学童", "夜間保育", 
        "施設形態", "施設区分", "病床数"
    ]
    
    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        # Optional: Make header bold
        from openpyxl.styles import Font
        cell.font = Font(bold=True)

    # Sort logic: North to South (JIS X 0401)
    prefectures = [
        "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
        "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
        "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
        "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
        "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
        "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
        "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県"
    ]
    
    # Create a rank map for faster lookup
    pref_rank = {p: i for i, p in enumerate(prefectures)}
    
    def get_sort_key(item):
        # Master record
        mas = item.get('master', {})
        
        # 1. Prefecture Rank
        # Extract addr_area directly
        addr = mas.get('addr_area', {}).get('value', "")
        rank = pref_rank.get(addr, 999)
        
        # 2. Municipality (City/Ward)
        city = mas.get('addr_city', {}).get('value', "") or ""
        
        # 3. Client Name
        client = mas.get('client_name', {}).get('value', "") or ""
        
        return (rank, city, client)

    # Sort merged_data in place with multi-key
    merged_data.sort(key=get_sort_key)

    # 2. Write Data
    # merged_data list of dicts: {'master': {...}, 'bed': {...}}
    
    row_idx = 2
    for i, item in enumerate(merged_data, 1):
        m = item.get('master', {})
        b = item.get('bed', {}) # Bed data might be array or single? Assuming 1-to-1 match logic from merge_data
        
        # Helper to safely get value
        def val(record, field):
            return record.get(field, {}).get('value', "")

        # Write Row
        # Col 1: Address (addr_area + addr_city)
        address = f"{val(m, 'addr_area')}{val(m, 'addr_city')}"
        ws.cell(row=row_idx, column=1).value = address
        
        ws.cell(row=row_idx, column=2).value = val(m, 'status')
        ws.cell(row=row_idx, column=3).value = val(m, 'name')
        ws.cell(row=row_idx, column=4).value = val(m, 'client_name')
        ws.cell(row=row_idx, column=5).value = val(m, 'open_date')
        
        # Checkbox for Basic Opening Days (likely a list)
        def fmt(v):
            if isinstance(v, list): return ", ".join(v)
            return v
            
        ws.cell(row=row_idx, column=6).value = fmt(val(m, '基本開園日'))
        ws.cell(row=row_idx, column=7).value = val(m, 'capacity')
        
        ws.cell(row=row_idx, column=8).value = fmt(val(m, 'sick_child_care'))
        ws.cell(row=row_idx, column=9).value = fmt(val(m, 'sc_flg'))
        ws.cell(row=row_idx, column=10).value = fmt(val(m, 'night_care'))
        ws.cell(row=row_idx, column=11).value = fmt(val(m, 'ekbn2'))
        ws.cell(row=row_idx, column=12).value = fmt(val(m, 'ekbn4'))
        
        # Bed count from bed app
        # bed_data is likely a list of records if multiple matched? 
        # Checking merge_data logic (not visible here but assuming structure)
        # Usually merge_data returns 'bed' as list of records matching the name.
        # Let's sum bed counts if multiple?
        bed_count = 0
        if isinstance(b, list):
            for brec in b:
                try:
                    bed_count += int(brec.get('病床数合計_0', {}).get('value', 0) or 0)
                except: pass
        elif isinstance(b, dict):
             try:
                bed_count = int(b.get('病床数合計_0', {}).get('value', 0) or 0)
             except: pass
             
        ws.cell(row=row_idx, column=13).value = bed_count
        
        row_idx += 1

    return wb
