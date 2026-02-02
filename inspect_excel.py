import openpyxl

wb = openpyxl.load_workbook('26.01.07_運営実績_2026年1月7日更新【kids営業】.xlsx')
print('Sheet names:', wb.sheetnames)
ws = wb.active
print('Active sheet:', ws.title)
print('Dimensions:', ws.dimensions)
print()

print('=== First 5 rows, first 15 columns ===')
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=15, values_only=True), 1):
    print(f"Row {row_idx}: {row}")

print()
print('=== Column headers (Row 1) ===')
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        headers.append((col, val))
print(f"Total columns with data: {len(headers)}")
for col, val in headers[:30]:  # First 30 headers
    print(f"  Col {col}: {val}")
